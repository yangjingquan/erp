from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from unicodedata import normalize as unicode_normalize

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.exceptions import AppError
from app.models.master_data import (
    MdCustomer,
    MdMaterial,
    MdSupplier,
    MdTaxRate,
    MdUnit,
    MdWarehouse,
)
from app.services.audit_service import write_operation_log

RESOURCE_CONFIG = {
    "materials": {
        "model": MdMaterial,
        "fields": [
            "code", "name", "category", "material_type", "standard_cost", "sale_price",
            "purchase_price", "min_stock", "max_stock", "specification",
        ],
    },
    "customers": {
        "model": MdCustomer,
        "fields": ["code", "name", "short_name", "owner_id", "contact_name", "contact_phone", "address", "credit_limit"],
    },
    "suppliers": {
        "model": MdSupplier,
        "fields": ["code", "name", "short_name", "owner_id", "contact_name", "contact_phone", "address", "credit_days"],
    },
    "warehouses": {"model": MdWarehouse, "fields": ["code", "name", "manager_id", "address"]},
    "units": {"model": MdUnit, "fields": ["code", "name", "precision_scale"]},
    "tax-rates": {"model": MdTaxRate, "fields": ["code", "name", "rate"]},
}


def normalize_name(value: str) -> str:
    return "".join(unicode_normalize("NFKC", value).split()).casefold()


def get_config(resource: str) -> dict[str, Any]:
    if resource not in RESOURCE_CONFIG:
        raise AppError("不支持的主数据类型", code=404)
    return RESOURCE_CONFIG[resource]


def _validate_resource_data(resource: str, data: dict[str, Any]) -> None:
    if resource != "tax-rates" or "rate" not in data:
        return
    try:
        rate = Decimal(str(data["rate"]))
    except (InvalidOperation, TypeError, ValueError):
        raise AppError("税率必须是 0 至 100 之间的数字", code=422)
    if rate < 0 or rate > 100:
        raise AppError("税率必须在 0 至 100 之间", code=422)


def list_items(db: Session, resource: str, org_id: str, keyword: str | None = None, page: int = 1, page_size: int = 200, code: str | None = None, name: str | None = None, category: str | None = None, material_type: str | None = None, status: str | None = None) -> tuple[list[Any], int, int, int]:
    model = get_config(resource)["model"]
    statement = select(model).where(model.org_id == org_id, model.is_deleted.is_(False))
    if keyword:
        pattern = f"%{keyword.strip()}%"
        statement = statement.where((model.code.like(pattern)) | (model.name.like(pattern)))
    for field, value in (("code", code), ("name", name), ("category", category), ("material_type", material_type), ("status", status)):
        if value and hasattr(model, field):
            statement = statement.where(getattr(model, field).like(f"%{value.strip()}%") if field != "status" else getattr(model, field) == value)
    total = db.scalar(select(func.count()).select_from(statement.subquery())) or 0
    active = db.scalar(select(func.count()).select_from(statement.where(model.status == "active").subquery())) if hasattr(model, "status") else total
    active = int(active or 0)
    total = int(total)
    page = max(page, 1)
    page_size = min(max(page_size, 1), 500)
    rows = list(db.scalars(statement.order_by(model.created_at.desc()).offset((page - 1) * page_size).limit(page_size)).all())
    return rows, total, active, total - active


def create_item(db: Session, resource: str, org_id: str, data: dict[str, Any], user: object):
    _validate_resource_data(resource, data)
    model = get_config(resource)["model"]
    existing = list(
        db.scalars(
            select(model).where(model.org_id == org_id, model.is_deleted.is_(False))
        ).all()
    )
    if any(row.code == data["code"] for row in existing):
        raise AppError("业务编码已存在", code=409)
    if any(normalize_name(row.name) == normalize_name(data["name"]) for row in existing):
        raise AppError("名称已存在", code=409)
    allowed_fields = set(get_config(resource)["fields"])
    instance = model(
        org_id=org_id,
        **{key: value for key, value in data.items() if key in allowed_fields},
    )
    db.add(instance)
    write_operation_log(
        db,
        user=user,
        action="create",
        resource=resource,
        target_id=instance.id,
        detail={"code": instance.code, "name": instance.name},
    )
    db.commit()
    db.refresh(instance)
    return instance


def update_item(db: Session, resource: str, item_id: str, org_id: str, data: dict[str, Any], user: object):
    _validate_resource_data(resource, data)
    model = get_config(resource)["model"]
    instance = db.scalar(select(model).where(model.id == item_id, model.org_id == org_id, model.is_deleted.is_(False)))
    if instance is None:
        raise AppError("主数据记录不存在", code=404)
    existing = db.scalars(select(model).where(model.org_id == org_id, model.is_deleted.is_(False), model.id != item_id)).all()
    if any(row.code == data["code"] for row in existing):
        raise AppError("业务编码已存在", code=409)
    if any(normalize_name(row.name) == normalize_name(data["name"]) for row in existing):
        raise AppError("名称已存在", code=409)
    for key in get_config(resource)["fields"]:
        if key in data:
            setattr(instance, key, data[key])
    instance.version += 1
    write_operation_log(db, user=user, action="update", resource=resource, target_id=instance.id, detail={"code": instance.code, "name": instance.name})
    db.commit()
    db.refresh(instance)
    return instance


def set_item_status(db: Session, resource: str, item_id: str, org_id: str, status: str, user: object):
    if status not in {"active", "inactive"}:
        raise AppError("状态无效", code=400)
    model = get_config(resource)["model"]
    instance = db.scalar(select(model).where(model.id == item_id, model.org_id == org_id, model.is_deleted.is_(False)))
    if instance is None:
        raise AppError("主数据记录不存在", code=404)
    if not hasattr(instance, "status"):
        raise AppError("该主数据类型不支持状态变更", code=400)
    if resource == "tax-rates" and status == "active" and not Decimal("0") <= instance.rate <= Decimal("100"):
        raise AppError("税率超出 0 至 100，修正后才能启用", code=422)
    instance.status = status
    instance.version += 1
    write_operation_log(db, user=user, action="status", resource=resource, target_id=instance.id, detail={"status": status})
    db.commit()
    db.refresh(instance)
    return instance


def serialize_item(item: Any, fields: list[str]) -> dict[str, Any]:
    result = {"id": item.id}
    if hasattr(item, "status"):
        result["status"] = item.status
    for field in fields:
        value = getattr(item, field, None)
        result[field] = str(value) if hasattr(value, "as_tuple") else value
    return result


def import_items(db: Session, resource: str, org_id: str, file_obj, user: object) -> dict:
    config = get_config(resource)
    model = config["model"]
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(value).strip() if value is not None else "" for value in next(sheet.iter_rows(values_only=True))]
    rows = []
    errors = []
    created_count = 0
    skipped_count = 0
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {headers[index]: values[index] for index in range(min(len(headers), len(values)))}
        code = str(row.get("code") or "").strip()
        name = str(row.get("name") or "").strip()
        if not code or not name:
            errors.append({"row": row_number, "message": "code 和 name 不能为空"})
            continue
        existing = list(db.scalars(select(model).where(model.org_id == org_id, model.is_deleted.is_(False))).all())
        if any(item.code == code or normalize_name(item.name) == normalize_name(name) for item in existing):
            skipped_count += 1
            continue
        data = {key: row.get(key) for key in config["fields"] if row.get(key) is not None}
        try:
            _validate_resource_data(resource, data)
            instance = model(org_id=org_id, **data)
            db.add(instance)
            db.flush()
            created_count += 1
        except Exception as exc:
            db.rollback()
            errors.append({"row": row_number, "message": str(exc)})
    db.commit()
    write_operation_log(
        db,
        user=user,
        action="import",
        resource=resource,
        detail={"created_count": created_count, "skipped_count": skipped_count, "errors": errors},
    )
    db.commit()
    return {"created_count": created_count, "skipped_count": skipped_count, "errors": errors}


def export_items(db: Session, resource: str, org_id: str) -> BytesIO:
    config = get_config(resource)
    workbook = Workbook()
    sheet = workbook.active
    fields = config["fields"]
    sheet.append(fields)
    for item in list_items(db, resource, org_id)[0]:
        sheet.append([getattr(item, field, None) for field in fields])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream
