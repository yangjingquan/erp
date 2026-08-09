from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.core.security import create_access_token


def auth_headers():
    return {"Authorization": f"Bearer {create_access_token('user-1', ['master:manage'])}"}


def test_create_material_and_reject_duplicate_code(client_and_session):
    client, _ = client_and_session
    payload = {"code": "MAT-001", "name": "螺丝", "sale_price": 2.5}

    first = client.post("/api/master/materials", json=payload, headers=auth_headers())
    duplicate = client.post("/api/master/materials", json=payload, headers=auth_headers())

    assert first.json()["code"] == 0
    assert duplicate.json()["code"] == 409


def test_customer_name_duplicate_ignores_spaces_and_width(client_and_session):
    client, _ = client_and_session
    headers = auth_headers()

    client.post(
        "/api/master/customers",
        json={"code": "CUS-001", "name": "客户 A"},
        headers=headers,
    )
    duplicate = client.post(
        "/api/master/customers",
        json={"code": "CUS-002", "name": "客户Ａ"},
        headers=headers,
    )

    assert duplicate.json()["code"] == 409


def test_excel_import_reports_valid_and_invalid_rows(client_and_session):
    client, _ = client_and_session
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["code", "name", "category"])
    sheet.append(["MAT-100", "进口轴承", "机械件"])
    sheet.append(["", "缺少编码", "机械件"])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    response = client.post(
        "/api/master/materials/import",
        files={
            "file": (
                "materials.xlsx",
                stream.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_headers(),
    )

    data = response.json()["data"]
    assert response.json()["code"] == 0
    assert data["created_count"] == 1
    assert data["errors"][0]["row"] == 3


def test_excel_export_returns_xlsx_content(client_and_session):
    client, _ = client_and_session
    client.post(
        "/api/master/materials",
        json={"code": "MAT-EXPORT", "name": "导出物料"},
        headers=auth_headers(),
    )

    response = client.get("/api/master/materials/export", headers=auth_headers())

    assert response.status_code == 200
    exported = load_workbook(BytesIO(response.content))
    assert exported.active["A1"].value == "code"
